import xlrd
import time
from openpyxl import workbook
from openpyxl import load_workbook

class Records:

    def get_inventory(self):
        workbook = xlrd.open_workbook('Records.xlsx')
        inventory = workbook.sheet_by_name('Inventory')

        row_count = inventory.nrows
        col_count = inventory.ncols

        print("Getting inventory...\n\n")
        time.sleep(1)

        keys = [inventory.cell(0, col_index).value for col_index in range(col_count)]

        dict_list = []
        for row_index in range(1, row_count):
            d = {keys[col_index]: inventory.cell(row_index, col_index).value
                 for col_index in range(col_count)}
            dict_list.append(d)
        return dict_list

    def display_inventory(self, dict_list):
        for i in range(len(dict_list)):
            print(dict_list[i])

    def search_inventory(self, dict_list):
        no_key = True
        no_value = True

        print("Choose from Headings:")
        for key in dict_list[0]:
            print(key)

        user_key = str(input("\nWhere: "))
        user_value = input("is: ")

        for dictionary in dict_list:
            for key in dictionary:
                if key == user_key:
                    no_key = False
                    if dictionary[key] == user_value:
                        print(dictionary)
                        no_value = False
                    elif user_value.isdigit():
                        if int(dictionary[key]) == int(user_value):
                            print(dictionary)
                            no_value = False
                    else:
                        try:
                            float_user_ip = float(user_value)
                            if dictionary[key] == float_user_ip:
                                print(dictionary)
                                no_value = False
                        except ValueError:
                            pass

        if no_value:
            if no_key:
                print("Key does not exist!")
            else:
                print("No value against this key!")

    def ammend_inventory(self, dict_list):
        id_ammend = input("ID Number to ammend: ")
        no_value = True
        key_id = "ID"
        row_index = 1
        col_index = 1
        for dictionary in dict_list:
            row_index = row_index + 1
            for key, value in dictionary.items():
                if key == key_id and value == id_ammend:
                    print(dictionary)
                    no_value = False
                    break
            else:
                continue
            break

        print("Row value is: " + str(row_index))  # This is the row
        if no_value:
                print("ID does not exist!")
        else:
            ammend_key = input("Heading of value to be amended: ")  # This is the col
            ammend_value = input("New value: ")  # This is the new val for [row, col]
            headings = dict_list[0].keys()
            for heading in headings:
                if heading != ammend_key:
                    col_index = col_index + 1
                else:
                    break

            print("Column value is: " + str(col_index))
            wb = load_workbook('Records.xlsx')
            ws = wb.get_sheet_by_name('Inventory')
            ws.cell(row=row_index, column=col_index).value = ammend_value
            wb.save('Records.xlsx')
            print("Value updated!")



    def vendors(self):
        workbook = xlrd.open_workbook('Records.xlsx')
        vendors = workbook.sheet_by_name('Vendors')
        print(vendors.cell(0, 1))

    def customers(self):
        workbook = xlrd.open_workbook('Records.xlsx')
        customers = workbook.sheet_by_name('Customers')
        print(customers.cell(0, 1))

    def accounts(self):
        workbook = xlrd.open_workbook('Records.xlsx')
        accounts = workbook.sheet_by_name('Accounts')
        print(accounts.cell(0, 1))